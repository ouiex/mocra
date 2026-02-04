"""Excel -> Polars DataFrame helpers (port of Rust utils::excel_dataframe)."""

from __future__ import annotations

import io
from typing import Any, List

from .type_convert import excel_values_to_series

try:
    import polars as pl
except Exception:  # pragma: no cover
    pl = None


def _require_polars():
    if pl is None:
        raise ImportError("polars is not installed")


def _sheet_dataframe(rows: List[List[Any]], first_row_as_title: bool):
    _require_polars()

    if not rows:
        return pl.DataFrame()

    columns: List[pl.Series] = []
    width = max(len(row) for row in rows)
    for col_index in range(width):
        col_values = [row[col_index] if col_index < len(row) else None for row in rows]
        if first_row_as_title:
            column_name = str(col_values[0]) if col_values else f"column_{col_index}"
            series = excel_values_to_series(column_name, col_values[1:])
        else:
            column_name = f"column_{col_index}"
            series = excel_values_to_series(column_name, col_values)
        columns.append(series)
    return pl.DataFrame(columns)


def xlsx_dataframe(data: bytes, sheet_name: str, skip_rows: int, first_row_as_title: bool):
    _require_polars()
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        raise ImportError("openpyxl is required for xlsx parsing") from exc

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the workbook")

    ws = wb[sheet_name]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < skip_rows:
            continue
        rows.append(list(row))
    return _sheet_dataframe(rows, first_row_as_title)


def xls_dataframe(data: bytes, sheet_name: str, skip_rows: int, first_row_as_title: bool):
    _require_polars()
    try:
        import xlrd  # type: ignore
    except Exception as exc:  # pragma: no cover
        raise ImportError("xlrd is required for xls parsing") from exc

    book = xlrd.open_workbook(file_contents=data)
    if sheet_name not in book.sheet_names():
        raise ValueError(f"Sheet '{sheet_name}' not found in the workbook")
    sheet = book.sheet_by_name(sheet_name)

    rows = []
    for row_idx in range(skip_rows, sheet.nrows):
        rows.append(sheet.row_values(row_idx))
    return _sheet_dataframe(rows, first_row_as_title)
